#Calculates payment schedule for Reg D Bonds, Compounding or Simple Interest
import pandas as pd
import calculations as c
import numpy as np
from datetime import datetime
from datetime import date
from pandas.plotting import table 
import os
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.datavalidation import DataValidation
import excel_sheet_formatting as ef 
from openpyxl.drawing.image import Image

def write_unique_file(output_path):
    
    base,ext = os.path.splitext(output_path)
    count = 0
    if os.path.exists(output_path):

        while os.path.exists(output_path):
                count += 1
                output_path = f'{base}_{count}{ext}'
        # print(output_path)
        return output_path
    else:
        # print(output_path)
        return output_path

def compound_interest_invest_value(initial, rate, term):
    #returns a list of cumulative interest values showing the total 
    # compounding interst and a list of the interest values paid for each month
    interest_value_by_month = 0
    investment_value = 0
    interst_val_each_month = []
    cumulative_interest_vals = []
    rateterm = ((rate/100)/12)

    for payment in range(term*12):
        
        interest_value_by_month = (initial + sum(interst_val_each_month))*(rateterm)
        investment_value += interest_value_by_month
        cumulative_interest_vals.append(investment_value)

        interst_val_each_month.append(interest_value_by_month)
    
    rounded_cum_interest_vals = []
    rounded_interst_vals_each_month = []
    
    for elem in interst_val_each_month:
        rounded_interst_vals_each_month.append(round(elem,2))

    for elem in cumulative_interest_vals:
        elem = elem + initial
        rounded_cum_interest_vals.append(float(round(elem,2)))

    return rounded_cum_interest_vals, rounded_interst_vals_each_month

def simple_interest_schedule_discounted(initial,rate,bond_yield,term):
    bond_price = c.bond_price(rate,term,bond_yield)
    adjusted_bond_total = c.adjusted_total_bonds(initial,bond_price)
    totalinterest = c.discount_total_interest(initial,term,adjusted_bond_total,rate)
    monthlyincome = c.discount_monthly_income(totalinterest,term)

    monthly_interest = []
    cumulative_interest_ = 0
    cumulative_interest_vals = []

    for payment in range(term*12):
        
        interest_value_by_month = round(monthlyincome,2)
        cumulative_interest_ += interest_value_by_month
        cumulative_interest_vals.append(cumulative_interest_)

        monthly_interest.append(interest_value_by_month)
    
    rounded_cum_interest_vals = []
    rounded_interst_vals_each_month = []
    
    for elem in monthly_interest:
        rounded_interst_vals_each_month.append(c.insert_commas_and_dollar(round(elem,2)))

    for elem in cumulative_interest_vals:
        rounded_cum_interest_vals.append(c.insert_commas_and_dollar(round(elem,2)))

    return rounded_cum_interest_vals, rounded_interst_vals_each_month 

def simple_interest_schedule(initial,rate,term):
    monthly_interest = []
    cumulative_interest_ = 0
    cumulative_interest_vals = []

    for payment in range(term*12):
        
        interest_value_by_month = round((initial*(rate/100))/12,2)
        cumulative_interest_ += interest_value_by_month
        cumulative_interest_vals.append(cumulative_interest_)

        monthly_interest.append(interest_value_by_month)
    
    rounded_cum_interest_vals = []
    rounded_interst_vals_each_month = []
    
    for elem in monthly_interest:
        rounded_interst_vals_each_month.append(c.insert_commas_and_dollar(round(elem,2)))

    for elem in cumulative_interest_vals:
        rounded_cum_interest_vals.append(c.insert_commas_and_dollar(round(elem,2)))

    return rounded_cum_interest_vals, rounded_interst_vals_each_month   

def compound_interest_schedule_discounted(initial, rate,bond_yield, term):
    #returns a list of cumulative interest values showing the total 
    # compounding interst and a list of the interest values paid for each month
    bond_price = c.bond_price(rate,term,bond_yield)
    adjusted_bond_total = c.adjusted_total_bonds(initial,bond_price)
    totalinterest = c.discount_total_interest(initial,term,adjusted_bond_total,rate)
    adjusted_initial = adjusted_bond_total * 1000
    
    interest_value_by_month = 0
    cumulative_interest_ = 0
    interst_val_each_month = []
    cumulative_interest_vals = []
    rateterm = ((rate/100)/12)

    for payment in range(term*12):
        
        interest_value_by_month = (adjusted_initial + sum(interst_val_each_month))*(rateterm)
        cumulative_interest_ += interest_value_by_month
        cumulative_interest_vals.append(cumulative_interest_)

        interst_val_each_month.append(interest_value_by_month)
    
    rounded_cum_interest_vals = []
    rounded_interst_vals_each_month = []
    
    for elem in interst_val_each_month:
        rounded_interst_vals_each_month.append(c.insert_commas_and_dollar(round(elem,2)))

    for elem in cumulative_interest_vals:
        rounded_cum_interest_vals.append(c.insert_commas_and_dollar(round(elem,2)))

    return rounded_cum_interest_vals, rounded_interst_vals_each_month

def compound_interest_schedule(initial, rate, term):
    #returns a list of cumulative interest values showing the total 
    # compounding interst and a list of the interest values paid for each month
    interest_value_by_month = 0
    cumulative_interest_ = 0
    interst_val_each_month = []
    cumulative_interest_vals = []
    rateterm = ((rate/100)/12)

    for payment in range(term*12):
        
        interest_value_by_month = (initial + sum(interst_val_each_month))*(rateterm)
        cumulative_interest_ += interest_value_by_month
        cumulative_interest_vals.append(cumulative_interest_)

        interst_val_each_month.append(interest_value_by_month)
    
    rounded_cum_interest_vals = []
    rounded_interst_vals_each_month = []
    
    for elem in interst_val_each_month:
        rounded_interst_vals_each_month.append(c.insert_commas_and_dollar(round(elem,2)))

    for elem in cumulative_interest_vals:
        rounded_cum_interest_vals.append(c.insert_commas_and_dollar(round(elem,2)))

    return rounded_cum_interest_vals, rounded_interst_vals_each_month

def generate_dates(settled_date, term):
    settled_date_object = datetime.strptime(settled_date, '%m/%d/%Y')

    settled_month = settled_date_object.month
    settled_year = settled_date_object.year
    settled_day = settled_date_object.day

    if settled_month < 12:
        first_payment_month = settled_month
        first_payment_year = settled_year
    else:
        first_payment_month = 12
        first_payment_year = settled_year

    first_payment_day = 10
    

    first_payment_date_object = datetime(first_payment_year, first_payment_month, first_payment_day)

    first_payment_date = datetime.date(first_payment_date_object)
    payment_dates = pd.date_range(first_payment_date_object, periods= term*12, freq='MS', inclusive='left') + pd.DateOffset(days=9)
    return payment_dates

def create_payment_table_compound(rounded_cum_interest_vals, rounded_interst_vals_each_month, dates):
    column_headers = ['Payment Dates', 'Accrued Interest', 'Cumulative Interest']
    payment_table = pd.DataFrame(data=[dates,rounded_interst_vals_each_month,rounded_cum_interest_vals], dtype='str').transpose()
    payment_table.columns = column_headers
    payment_table['Payment Dates'] = pd.to_datetime(payment_table['Payment Dates']).dt.normalize()
    
    return payment_table

def create_payment_table_simple(cumulative_interest_vals, dates, monthly_interest):
    column_headers = ['Payment Dates', 'Monthly Interest', 'Cumulative Interest']
    payment_table = pd.DataFrame(data=[dates, monthly_interest, cumulative_interest_vals], dtype='str').transpose()
    payment_table.columns = column_headers
    payment_table['Payment Dates'] = pd.to_datetime(payment_table['Payment Dates']).dt.normalize()
    
    return payment_table

def sum_years_c(df):
    df['Payment Dates']= pd.to_datetime(df['Payment Dates'])
    df['Accrued Interest'] = df['Accrued Interest'].replace('[\$,]', '', regex=True).astype(float)
    yearly_totals = df.resample('Y',on='Payment Dates')['Accrued Interest'].sum()
    yearly_counts = df.resample('Y',on='Payment Dates')['Accrued Interest'].count()

    yearly_totals_list = [''] * df.shape[0]

    indices = []
    last = 0
    for val in yearly_counts:
        indices.append((val+last)-1)
        last += val  

    i = 0
    for x in indices:
        yearly_totals_list[x] = c.insert_commas_and_dollar(yearly_totals[i])

        i +=1

    df['Yearly Totals'] = yearly_totals_list
    # df['Yearly Totals'] = df['Yearly Totals'].map('${:,.2f}'.format)
    df['Accrued Interest'] = df['Accrued Interest'].map('${:,.2f}'.format)

    return df     
 
def sum_years_s(df):
    df['Payment Dates']= pd.to_datetime(df['Payment Dates'])
    df['Monthly Interest'] = df['Monthly Interest'].replace('[\$,]', '', regex=True).astype(float)
    yearly_totals = df.resample('Y',on='Payment Dates')['Monthly Interest'].sum()
    yearly_counts = df.resample('Y',on='Payment Dates')['Monthly Interest'].count()

    yearly_totals_list = [''] * df.shape[0]

    indices = []
    last = 0
    for val in yearly_counts:
        indices.append((val+last)-1)
        last += val  

    i = 0
    for x in indices:
        yearly_totals_list[x] = c.insert_commas_and_dollar(yearly_totals[i])

        i +=1

    df['Yearly Totals'] = yearly_totals_list
    # df['Yearly Totals'] = df['Yearly Totals'].map('${:,.2f}'.format)
    df['Monthly Interest'] = df['Monthly Interest'].map('${:,.2f}'.format)

    return df  
    
def save_term_sheet_as_xlsx_compound(comp_data, comp_mature, initial, term, rate, filepath):
    #format data frame
    comp_data['Payment Dates'] = comp_data['Payment Dates'].dt.strftime("%b %d, %Y")
    comp_data.loc[len(comp_data.index)] = ['Initial Investment', ' ' , initial, ' ']
    comp_data.loc[len(comp_data.index)] = ['Mature Value', ' ' , comp_mature, ' ']



    #write to file
    with pd.ExcelWriter(filepath, engine='xlsxwriter') as writer:
        comp_data.to_excel(writer, index = False, sheet_name ='Compound Schedule', startrow = 1)


    #format excel file
        
        workbook = writer.book
        compound_sheet = writer.sheets['Compound Schedule']
        #insert image
        # image_path = "Logo Final - No text.png" 
        # png = openpyxl.drawing.image.Image(image_path)
        # png.height = .35
        # png.width = .35
        # png.anchor = compound_sheet['A1']
        # compound_sheet.add_image(image_path)
        # png.anchor = simple_sheet['A1']
        # simple_sheet.add_image(image_path)

        image_path = "Logo Final - No text.png" 
        compound_sheet.insert_image('A1', image_path, {'x_scale': .35, 'y_scale': .35})

#formatting 
        header_format = workbook.add_format({"valign": "vcenter",
        "align": "center",
        "bg_color": "#000080",
         "bold": True,
        'font_color': '#FFFFFF',
         'border' : 1, 
         'border_color': ''#D3D3D3'
         })
        comp_title = 'Reg D Compounding Interest Schedule:' + ' ' + str(term) + ' Year ' + str(rate) + '%'
        simple_title = 'Reg D Simple Interest Schedule:' + ' ' + str(term) + ' Year ' + str(rate) + '%'

        format_title = workbook.add_format()
        format_title.set_font_size(16)
        format_title.set_bg_color("#000080")
        format_title.set_font_color("#FFFFFF")
        format_title.set_align('center')
        format_title.set_bold(True)

        compound_sheet.merge_range('A1:D1', comp_title, format_title)


        compound_sheet.set_row(1, 22)
        compound_sheet.set_row(2,18)


        clean_sheet = workbook.add_format({'border':0})
        compound_sheet.set_column(4,18, 20, clean_sheet)


        format_data = workbook.add_format({"valign": "vcenter",
        "align": "center",
        'font_color': '#000000'})

        format_lastrows = workbook.add_format({"valign": "vcenter",
        "align": "center",
        'font_color': '#000000',
         'bold':True})
        
        rangeend = comp_data.shape[0] + 2
        for row in range(2,rangeend):
            if row <= rangeend - 3:
                compound_sheet.set_row(row,14, format_data)

            elif row in (rangeend-1, rangeend-2):
                compound_sheet.set_row(row,16, format_lastrows)

        for row in range(500):    
            if row >= rangeend:
                compound_sheet.set_row(row, 10, clean_sheet)

        for col in range(11):
            if col <= 2:
                compound_sheet.set_column(0,3,25, format_data)

            elif col > 2:
                clean_sheet.set_border(0)
                compound_sheet.set_column(4,18, 20, clean_sheet)


        for col_num, value in enumerate(comp_data.columns.values):
            compound_sheet.write(1, col_num, value, header_format)

#change range with size of dataframe
        end_of_data_sheet_value = comp_data.shape[0]+2
        conditional_range_border = 'A1:D'+ str(end_of_data_sheet_value)

        border_format = workbook.add_format({'border':1})
        compound_sheet.conditional_format( conditional_range_border , { 'type' : 'no_blanks' , 'format' : border_format})
        compound_sheet.conditional_format( conditional_range_border , { 'type' : 'blanks' , 'format' : border_format})


        conditional_range_last_2rows = 'A' + str(end_of_data_sheet_value-1) + ':D' + str(end_of_data_sheet_value)
        highlight_format = workbook.add_format({"bg_color": "#D9D9D9"})
        compound_sheet.conditional_format( conditional_range_last_2rows , { 'type' : 'no_blanks' , 'format' : highlight_format})
        compound_sheet.conditional_format( conditional_range_last_2rows , { 'type' : 'blanks' , 'format' : highlight_format})


    return filepath   

def save_term_sheet_as_xlsx_simple(simple_data, simple_mature, initial, term, rate, filepath):
    #format data frame

    simple_data['Payment Dates'] = simple_data['Payment Dates'].dt.strftime("%b %d, %Y")
    simple_data.loc[len(simple_data.index)] = ['Initial Investment', ' ' , initial, ' ']
    simple_data.loc[len(simple_data.index)] = ['Mature Value', ' ' , simple_mature, ' ']  

    #write to file
    with pd.ExcelWriter(filepath, engine='xlsxwriter') as writer:
        simple_data.to_excel(writer, index=False, sheet_name='Simple Schedule', startrow=1)

    #format excel file
        
        workbook = writer.book
        simple_sheet = writer.sheets['Simple Schedule']
        #insert image
        # image_path = "Logo Final - No text.png" 
        # png = openpyxl.drawing.image.Image(image_path)
        # png.height = .35
        # png.width = .35
        # png.anchor = compound_sheet['A1']
        # compound_sheet.add_image(image_path)
        # png.anchor = simple_sheet['A1']
        # simple_sheet.add_image(image_path)

        image_path = "Logo Final - No text.png" 

        simple_sheet.insert_image('A1', image_path, {'x_scale': .35, 'y_scale': .35})
#formatting 
        header_format = workbook.add_format({"valign": "vcenter",
        "align": "center",
        "bg_color": "#000080",
         "bold": True,
        'font_color': '#FFFFFF',
         'border' : 1, 
         'border_color': ''#D3D3D3'
         })
        comp_title = 'Reg D Compounding Interest Schedule:' + ' ' + str(term) + ' Year ' + str(rate) + '%'
        simple_title = 'Reg D Simple Interest Schedule:' + ' ' + str(term) + ' Year ' + str(rate) + '%'

        format_title = workbook.add_format()
        format_title.set_font_size(16)
        format_title.set_bg_color("#000080")
        format_title.set_font_color("#FFFFFF")
        format_title.set_align('center')
        format_title.set_bold(True)


        simple_sheet.merge_range('A1:D1', simple_title, format_title)


        simple_sheet.set_row(1, 22)
        simple_sheet.set_row(2,18)

        clean_sheet = workbook.add_format({'border':0})
        simple_sheet.set_column(4,18, 20, clean_sheet)

        format_data = workbook.add_format({"valign": "vcenter",
        "align": "center",
        'font_color': '#000000'})

        format_lastrows = workbook.add_format({"valign": "vcenter",
        "align": "center",
        'font_color': '#000000',
         'bold':True})
        
        rangeend = simple_data.shape[0] + 2
        for row in range(2,rangeend):
            if row <= rangeend - 3:

                simple_sheet.set_row(row,14, format_data)
            elif row in (rangeend-1, rangeend-2):

                simple_sheet.set_row(row,16, format_lastrows)
        for row in range(500):    
            if row >= rangeend:

                simple_sheet.set_row(row, 10, clean_sheet)
        for col in range(11):
            if col <= 2:

                simple_sheet.set_column(0,3,25, format_data)
            elif col > 2:
                clean_sheet.set_border(0)

                simple_sheet.set_column(4,18, 20, clean_sheet)

        for col_num, value in enumerate(simple_data.columns.values):
            simple_sheet.write(1, col_num, value, header_format)

#change range with size of dataframe
        end_of_data_sheet_value = simple_data.shape[0]+2
        conditional_range_border = 'A1:D'+ str(end_of_data_sheet_value)

        border_format = workbook.add_format({'border':1})

        simple_sheet.conditional_format( conditional_range_border , { 'type' : 'no_blanks' , 'format' : border_format})
        simple_sheet.conditional_format( conditional_range_border , { 'type' : 'blanks' , 'format' : border_format})

        conditional_range_last_2rows = 'A' + str(end_of_data_sheet_value-1) + ':D' + str(end_of_data_sheet_value)
        highlight_format = workbook.add_format({"bg_color": "#D9D9D9"})

        simple_sheet.conditional_format( conditional_range_last_2rows , { 'type' : 'no_blanks' , 'format' : highlight_format})
        simple_sheet.conditional_format( conditional_range_last_2rows , { 'type' : 'blanks' , 'format' : highlight_format})

    return filepath
 
def write_to_excel(data,initial, mature, term,settled_date, interest_type,investor_name,investor_email,discountTF,rate_lift,ymax_compare,ymax_discount,out_filepath, template_path):
    decimal_rate_lift = float(rate_lift)/100

    #define rate term
    if discountTF == False:
        rate = rate_lift
        titlerate = rate_lift
    else:
        if term == 1:
            rate = 9
        elif term == 3:
            rate = 10
        elif term == 5:
            rate = 11
        elif term ==7:
            rate = 12
        elif term == 9:
            rate = 12.5
        elif term == 11:
            rate = 13
        titlerate = rate
        
        mature = c.insert_commas_and_dollar(mature)
    maturity = data['Payment Dates'].iloc[-1]
    maturity_date = maturity.strftime('%b %d, %Y')

    if interest_type == "Simple":
        schedule_title_1 = 'Simple Interest Payment Schedule'
        schedule_title_2 = str(term) + ' Year at ' + str(titlerate) + '%'
        data['Payment Dates'] = data['Payment Dates'].dt.strftime("%b %d, %Y")
        data.loc[-1] = ['Initial Investment', ' ' , ' ', c.insert_commas_and_dollar(initial)]
        data.index = data.index + 1
        data = data.sort_index()
        data.loc[len(data.index)] = ['Mature Value', ' ' , ' ', mature]  

    elif interest_type == "Compounding":
        schedule_title_1 = 'Compounding Interest Payment Schedule'
        schedule_title_2 = str(term) + ' Year at ' + str(titlerate) + '%'
        data['Payment Dates'] = data['Payment Dates'].dt.strftime("%b %d, %Y")
        data.loc[-1] = ['Initial Investment', ' ' , ' ', c.insert_commas_and_dollar(initial)]
        data.index = data.index + 1
        data = data.sort_index()        
        data.loc[len(data.index)] = ['Mature Value', ' ' , ' ', mature]        

    writer = pd.ExcelWriter(template_path,engine='openpyxl', mode='a', if_sheet_exists='overlay', keep_vba=True)
    workbook = writer.book    

    comparison_sheet1 = workbook['Compare Offerings']
    drilldown_sheet1 = workbook['Drill Down']
    # workbook.create_sheet(title='Payment Schedule')

#data validation for both sheets drill down and compare offerings
    data_val_interest = DataValidation(type='list', formula1="='Data - Drill Down'!$B$28:$B$29")
    comparison_sheet1.add_data_validation(data_val_interest)
    data_val_interest.add(comparison_sheet1['B10'])
    drilldown_sheet1.add_data_validation(data_val_interest)
    data_val_interest.add(drilldown_sheet1['B10'])

    data_val_acct = DataValidation(type='list', formula1="='Data - Drill Down'!$B$10:$B$25")
    comparison_sheet1.add_data_validation(data_val_acct)
    data_val_acct.add(comparison_sheet1['B13'])
    drilldown_sheet1.add_data_validation(data_val_acct)
    data_val_acct.add(drilldown_sheet1['B13'])


    # data.to_excel(writer, index=False, sheet_name = "Payment Schedule")
    workbook.save(out_filepath)

    workbook = openpyxl.load_workbook(out_filepath, keep_vba=True)

    comparison_sheet = workbook['Compare Offerings']
    drilldown_sheet = workbook['Drill Down']   
    schedule_sheet = workbook['Payment Schedule']

    ef.format_excel_sheet(schedule_sheet, data)

    for row_num, row_data in enumerate(data.itertuples(), 4):  # Start from the second row
        for col_num, value in enumerate(row_data[1:], 1):  # Skip the index in row_data
            if col_num == 4 and value != '':
                for i in range(1,5):
                    schedule_sheet.cell(row=row_num, column=i).fill = PatternFill("solid", fgColor="00BFBFBF")
                    schedule_sheet.cell(row=row_num, column=i).font = Font(bold=True, color="00000000")
            else:
                continue

    #write investment inputs to sheets
    date_og = datetime.strptime(settled_date,  '%m/%d/%Y')
    settled_date_towrite = date_og.strftime('%b %d, %Y')

    comparison_sheet['B7'] = initial
    comparison_sheet['B8'] = term
    comparison_sheet['B10'] = interest_type
    comparison_sheet['B11'] = investor_name
    comparison_sheet['B12'] = investor_email
    comparison_sheet['F11'] = settled_date_towrite
    comparison_sheet['F13'] = maturity_date

    drilldown_sheet['B8'] = term
    drilldown_sheet['B10'] = interest_type
    drilldown_sheet['B11'] = investor_name
    drilldown_sheet['B12'] = investor_email
    drilldown_sheet['F11'] = settled_date_towrite
    drilldown_sheet['F13'] = maturity_date


    # chart_comparesheet = comparison_sheet._charts
    # chart_comparesheet[0].y_axis.scaling.max = ymax_compare

    # chart_drillsheet = drilldown_sheet._charts
    # chart_drillsheet[0].y_axis.scaling.max = ymax_discount
    #write discount inputs to drill down sheet

    drilldown_sheet['B25'] = decimal_rate_lift
    drilldown_sheet['B25'].number_format = '0.00%'

    highlight_border = Side(border_style="thick", color="00FF0000")
    for row_num in range(17,22):
        if term == 1:
            drilldown_sheet.cell(row=row_num, column=6).border = Border(top=highlight_border, left= highlight_border, right= highlight_border, bottom=highlight_border)
        elif term == 3:
            drilldown_sheet.cell(row=row_num, column=7).border = Border(top=highlight_border, left= highlight_border, right= highlight_border, bottom=highlight_border)
        elif term == 5:
            drilldown_sheet.cell(row=row_num, column=8).border = Border(top=highlight_border, left= highlight_border, right= highlight_border, bottom=highlight_border)
        elif term ==7:
            drilldown_sheet.cell(row=row_num, column=9).border = Border(top=highlight_border, left= highlight_border, right= highlight_border, bottom=highlight_border)
        elif term == 9:
            drilldown_sheet.cell(row=row_num, column=10).border = Border(top=highlight_border, left= highlight_border, right= highlight_border, bottom=highlight_border)
        elif term == 11:
            drilldown_sheet.cell(row=row_num, column=11).border = Border(top=highlight_border, left= highlight_border, right= highlight_border, bottom=highlight_border)

    for row_num in range(23,25):
        if term == 1:
            drilldown_sheet.cell(row=row_num, column=6).border = Border(top=highlight_border, left= highlight_border, right= highlight_border, bottom=highlight_border)
        elif term == 3:
            drilldown_sheet.cell(row=row_num, column=7).border = Border(top=highlight_border, left= highlight_border, right= highlight_border, bottom=highlight_border)
        elif term == 5:
            drilldown_sheet.cell(row=row_num, column=8).border = Border(top=highlight_border, left= highlight_border, right= highlight_border, bottom=highlight_border)
        elif term ==7:
            drilldown_sheet.cell(row=row_num, column=9).border = Border(top=highlight_border, left= highlight_border, right= highlight_border, bottom=highlight_border)
        elif term == 9:
            drilldown_sheet.cell(row=row_num, column=10).border = Border(top=highlight_border, left= highlight_border, right= highlight_border, bottom=highlight_border)
        elif term == 11:
            drilldown_sheet.cell(row=row_num, column=11).border = Border(top=highlight_border, left= highlight_border, right= highlight_border, bottom=highlight_border)

    #write to payment schedule and format
    schedule_sheet['A1'] = schedule_title_1
    schedule_sheet['A2'] = schedule_title_2
    # schedule_sheet['A1'].font = Font(bold=True, color='00FFFFFF', size=18)
    # schedule_sheet['A1'].fill = PatternFill("solid", fgColor="00003366")
    # schedule_sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    # schedule_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    # schedule_sheet.row_dimensions[1].height = 67.5
    # schedule_sheet.column_dimensions['A'].width = 36
    # schedule_sheet.column_dimensions['D'].width = 36


    # logo = Image("Logo Final - No text.png")
    # logo.height = 116
    # logo.width = 111
    # schedule_sheet.add_image(logo, 'A1')
    # schedule_sheet.add_image(logo, 'D1')  
    # logo2 = Image("Logo Final - No text.png")
    # logo2.height = 116
    # logo2.width = 111
    # schedule_sheet.add_image(logo2, 'D1')

    df_length = data.shape[0]
    df_width = data.shape[1]

    thick = Side(border_style="thick", color="000000")
    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="000000")

    for col in range(1,5):
        # schedule_sheet.cell(row=df_length+1, column=col).fill = PatternFill("solid", fgColor="00003366")
        schedule_sheet.cell(row=df_length+3, column=col).fill = PatternFill("solid", fgColor="00BF8F00")
        schedule_sheet.cell(row=4, column=col).fill = PatternFill("solid", fgColor="00BF8F00")
        schedule_sheet.cell(row=4, column=col).font = Font(bold=True, color='00000000', size=12)
        schedule_sheet.cell(row=df_length+2, column=col).border = Border(top=thin, left=thin, right=thin, bottom=thick)
        schedule_sheet.cell(row=df_length+3, column=col).border = Border(top=thick, bottom=thick)
        schedule_sheet.cell(row=df_length+2, column=col).font = Font(bold=True, color='00000000')
        schedule_sheet.cell(row=df_length+3, column=col).font = Font(bold=True, color='00000000', size=12)

    schedule_sheet.cell(row=df_length+3, column=4).border = Border(top=thick, bottom=thick, right=thin)
    schedule_sheet.cell(row=df_length+3, column=1).border = Border(top=thick, bottom=thick, left=thin)
    schedule_sheet.cell(row=3, column=4).border = Border(top=thick, bottom=thick,right=thin)
    schedule_sheet.cell(row=3, column=1).border = Border(top=thick, bottom=thick, left=thin)
    schedule_sheet.cell(row=4, column=4).border = Border(top=thick, bottom=thick,right=thin)
    schedule_sheet.cell(row=4, column=1).border = Border(top=thick, bottom=thick,left=thin)

    workbook.save(out_filepath)


# payments_s = simple_interest_schedule(50000,13,11)
# payments_c = compound_interest_schedule(1000000,14,7)
# df_comp = create_payment_table_compound(payments_c[0],payments_c[1],generate_dates('2/08/2024',7))
# # # # # # # # print(payments)
# # dfsimp = create_payment_table_simple(payments_s[0],generate_dates('1/05/2024',11), payments_s[1])
# data = sum_years_c(df_comp)
# # # print(dfsimp['Payment Dates'].iloc[-1])
# # # # # print(df_comp)
# # data = sum_years_s(dfsimp)
# # # print(data)

# todays_date = str(date.today())
# name = 'Term Sheet marco test'  + todays_date + '.xlsm'

# output_path = "C:\\Users\\Hannah Nevel\\Documents\\Interest Calculator\\12-13-2023 Scripts V3\\Testing\\" + name

# # unique_outputpath = write_unique_file(output_path)
# template_path = os.getcwd() + '\\Revenue Statement Template.xlsm'

# write_to_excel(data,1000000,'$2,649,384.66',7,'2/08/2024','Compounding','','',False,14,1400000,1400000,output_path,template_path)

