import os
import pandas as pd
import numpy as np
from datetime import datetime
from datetime import date
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

#FUNCTIONS NOT IN USE

def acct_name_with_LEAD(filepath):
    
    contacts_sheet = pd.read_excel(filepath, sheet_name='Investor Accounts with Contacts', na_values='')
    investor_accts_contacts = pd.DataFrame(contacts_sheet)
    
    with_lead = []
    
    for i,row in investor_accts_contacts.iterrows():
        if ' - LEAD' not in row['Account Name']:
            new_name = row['Account Name'] + ' - LEAD'
            with_lead.append(new_name)
        else:
            with_lead.append(row['Account Name'])

    investor_accts_contacts['Account Name with LEAD'] = with_lead
    investor_accts_contacts['Matchcode'] = investor_accts_contacts['Account ID'].astype(str) + "-" + investor_accts_contacts['Account Name with LEAD'].astype(str)
    col_names = ['Matchcode','Account Name with LEAD', 'Account ID', 'Account Name', 'Investment Owner', 'Investment Disposition', 'Confirmed Email Addresses', 'Last Name', 'Phone', 'Email', 'Force Sync', 'Contact ID']
    investor_accts_contacts = investor_accts_contacts[col_names]
    
    with pd.ExcelWriter(filepath, engine='openpyxl', mode='a') as writer:
        workbook = writer.book
        del workbook['Investor Accounts with Contacts']
        investor_accts_contacts.to_excel(writer, index=False, sheet_name='Investor Accounts with Contacts')

        worksheet = workbook['Investor Accounts with Contacts']
        worksheet.cell(row=1,column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        worksheet.cell(row=1,column=2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
def acct_name_with_LEAD_faster_2(filepath):
    
    contacts_sheet = pd.read_excel(filepath, sheet_name='Investor Accounts with Contacts', na_values='')
    investor_accts_contacts = pd.DataFrame(contacts_sheet)
    
    investor_accts_contacts['Account Name with LEAD'] = [investor_accts_contacts['Account Name'] if ' - LEAD' in x else investor_accts_contacts['Account Name']+' - LEAD' for x in investor_accts_contacts['Account Name']]
    investor_accts_contacts['Matchcode'] = investor_accts_contacts['Account ID'].astype(str) + "-" + investor_accts_contacts['Account Name with LEAD'].astype(str)
    col_names = ['Matchcode','Account Name with LEAD', 'Account ID', 'Account Name', 'Investment Owner', 'Investment Disposition', 'Confirmed Email Addresses', 'Last Name', 'Phone', 'Email', 'Force Sync', 'Contact ID']
    investor_accts_contacts = investor_accts_contacts[col_names]

    with pd.ExcelWriter(filepath, engine='openpyxl', mode='a') as writer:
        workbook = writer.book
        del workbook['Investor Accounts with Contacts']
        investor_accts_contacts.to_excel(writer, index=False, sheet_name='Investor Accounts with Contacts')

        worksheet = workbook['Investor Accounts with Contacts']
        worksheet.cell(row=1,column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        worksheet.cell(row=1,column=2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")