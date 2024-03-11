import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# Function to format the Excel sheet
def format_excel_sheet(sheet, df):
    bd = Side(border_style='thick')
    border_header = Border(top=bd, bottom=bd)
    thin = Side(border_style="thin", color="000000")
    border_dataset = Border(top=thin,bottom=thin,right=thin,left=thin)
   

# Write headers
    for col_num, header in enumerate(df.columns, 1):
        sheet.cell(row=3, column=col_num, value=header).font = Font(bold=True, color='00000000', size=14)
        sheet.cell(row=3, column=col_num).alignment = Alignment(horizontal='center')
        sheet.cell(row=3, column=col_num).border = border_header
        sheet.cell(row=3, column=col_num).fill = PatternFill("solid", fgColor="00BF8F00")

    # Write data
    for row_num, row_data in enumerate(df.itertuples(), 4):  # Start from the fourth row
        for col_num, value in enumerate(row_data[1:], 1):  # Skip the index in row_data
            if row_num ==4:
                sheet.cell(row=row_num, column=col_num, value = value).alignment = Alignment(horizontal='center')
                sheet.cell(row=row_num, column =col_num, value=value).font = Font(bold=True,color='00000000', size =12)
                sheet.cell(row=row_num, column=col_num).border = border_header
            else:

                sheet.cell(row=row_num, column=col_num, value=value).alignment = Alignment(horizontal='center')
                sheet.cell(row=row_num, column=col_num, value=value).border = border_dataset

    # Auto-size columns
    # for col in sheet.columns:
    #     max_length = 0
    #     column = col[0].column_letter
    #     for cell in col:
    #         try:
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(cell.value)
    #         except:
    #             pass
    #     adjusted_width = (max_length + 2)
    #     sheet.column_dimensions[column].width = adjusted_width
