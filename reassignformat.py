#reassign formatting 
import os
import pandas as pd
import numpy as np
from datetime import datetime
from datetime import date
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from tqdm import tqdm
import PySimpleGUI as sg

def import_xlsxfiles(file):
    #import datafile from local folder, covert to dataframe for manipulation
    # path = os.getcwd()+'\\'+filename
    readfile = pd.read_excel(file, na_values='', dtype='str', engine="openpyxl")
    df = pd.DataFrame(readfile)

    return df

def import_csvfiles(file):
    #import datafile from local folder, covert to dataframe for manipulation
    # path = os.getcwd()+'\\'+filename
    readfile = pd.read_csv(file, encoding='ISO-8859-1', na_values='', dtype='str')
    df = pd.DataFrame(readfile)

    return df

def create_path(datafile_path):
    data_filename = os.path.basename(datafile_path)
    directory = datafile_path.replace(data_filename,'')
    directory_1 = directory[:-1].replace('/','\\\\')
    adjusted_datafile_path = directory_1 + '\\\\' + data_filename

    return adjusted_datafile_path

def create_folder(outputpath, name):
    todays_date = str(date.today()).replace("-","")
    foldername = todays_date + name
    base,ext = os.path.splitext(foldername)
    count = 0

    while True:
        try:
            output_filepath = outputpath + '\\' + foldername
    
            os.makedirs(output_filepath, exist_ok=False)
        except OSError:
            count += 1
            foldername = f'{base}_{count}{ext}'
        else:
            return output_filepath


def save_each_file_to_folder(calllist,contacts,activities,all_invest,output_path):
    calllist.to_excel((output_path+'\\'+'New Investor Call list to distribute.xlsx'),index=False)
    contacts.to_csv((output_path+'\\'+'Investor Accounts with Contacts.csv'),index=False)
    activities.to_csv((output_path+'\\'+'Investor Activities Webinar Session.csv'),index=False)
    all_invest.to_csv((output_path+'\\'+'All Investment Accounts.csv'),index=False)

def filter_future_web(reassigndf,futureweb,pastweb):
    reassigndf['Future Web Check'] = reassigndf['Confirmed Email Addresses'].str.lower().isin(futureweb['Email'].str.lower())
    reassigndf = reassigndf[reassigndf['Future Web Check'] == False]

    reassigndf['attended check'] = reassigndf['Confirmed Email Addresses'].str.lower().isin(pastweb['Email'].str.lower())
    
    return reassigndf

def combine_files_webinar(calllist,contacts,activities,all_invest,output_path):
    accounts_to_reassign = import_xlsxfiles(calllist)
    investor_accts_contacts = import_csvfiles(contacts)
    investor_activities_web = import_csvfiles(activities)
    all_investment_accts = import_csvfiles(all_invest)

    save_each_file_to_folder(accounts_to_reassign,investor_accts_contacts,investor_activities_web,all_investment_accts,output_path)

    working_df = accounts_to_reassign.copy()

    todays_date = str(date.today())
    filename = todays_date + ' - Investor Reassign Call List.xlsx'
    output_filepath = output_path + '\\' + filename

    user_path = os.getcwd()+'\\User Table - 20231026.xlsx'
    user_table = import_xlsxfiles(user_path)

    contacts = acct_name_with_LEAD_faster(investor_accts_contacts)


    split_web_dates = past_future_webinar(investor_activities_web)
    past_webinar = split_web_dates[0]
    future_webinar = split_web_dates[1]

    modified_accts_to_reassign = insert_reassign_columns(working_df)

    modified_accts_to_reassign = filter_future_web(modified_accts_to_reassign,future_webinar,past_webinar)

   
    with pd.ExcelWriter(output_filepath, engine='openpyxl') as writer:
        accounts_to_reassign.to_excel(writer, index = False, sheet_name ='NewInvestors-CallListToDistribu')
        modified_accts_to_reassign.to_excel(writer, index = False, sheet_name ='NewInvestors-CallListToDist (2)')
        investor_activities_web.to_excel(writer, index=False, sheet_name='Investor Activities Webinar Ses')
        past_webinar.to_excel(writer, index=False, sheet_name='Past Webinar')
        future_webinar.to_excel(writer, index=False, sheet_name='Future Webinar')
        all_investment_accts.to_excel(writer, index=False, sheet_name='All Investment Accounts')
        contacts.to_excel(writer, index=False, sheet_name='Investor Accounts with Contacts')
        user_table.to_excel(writer, index=False, sheet_name='User')

        workbook = writer.book

        worksheet_contacts = workbook['Investor Accounts with Contacts']
        worksheet_contacts.cell(row=1,column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        worksheet_contacts.cell(row=1,column=2).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        worksheet_reassign = workbook['NewInvestors-CallListToDist (2)']

        no_phone_sheet = workbook.create_sheet(title='No Phone')

        for i in range(2,len(modified_accts_to_reassign)+2):
            worksheet_reassign['Q'+str(i)] = "=INDEX('Future Webinar'!G:G,MATCH(J"+ str(i) +",'Future Webinar'!I:I,0))"
            worksheet_reassign['R'+str(i)] =  "=INDEX('Past Webinar'!K:K,MATCH(J"+ str(i) +",'Past Webinar'!I:I,0))"
            worksheet_reassign['S'+str(i)] = "=INDEX('Future Webinar'!H:H,MATCH(J"+ str(i) +",'Future Webinar'!I:I,0))"
            worksheet_reassign['T'+str(i)] = "=INDEX('Past Webinar'!H:H,MATCH(J"+ str(i) +",'Past Webinar'!I:I,0))"
            worksheet_reassign['E'+str(i)] = "=IF(VLOOKUP(C"+ str(i) + ", 'Investor Accounts with Contacts'!A:L, 11, FALSE)=0, TRUE, FALSE)"
            worksheet_reassign['F'+str(i)] = "=VLOOKUP(C"+ str(i) +",'Investor Accounts with Contacts'!A:L,12,FALSE)"

        blue = Font(color='000000FF')
        for d,i in zip(worksheet_reassign['D'], worksheet_reassign['I']):
            if d.row == 1:
                continue
            if d.value != i.value:
                d.font = blue

        for a,c in zip(worksheet_reassign['A'], worksheet_reassign['C']):
            if c.row ==1:
                continue
            else:
                c.font = blue 

        for b,d in zip(worksheet_contacts['B'], worksheet_contacts['D']):
            if b.row ==1:
                continue
            if b.value != d.value:
                b.font = blue

        for a,b in zip(worksheet_contacts['A'], worksheet_contacts['B']):
            if a.row ==1:
                continue
            else:
                a.font = blue

        for i in range(1,worksheet_reassign.max_column+1):
            c = worksheet_reassign.cell(row=1, column=i)
            no_phone_sheet.cell(row=1, column=i).value = c.value

        for i in range(1,7):
            worksheet_reassign.cell(row=1,column=i).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            no_phone_sheet.cell(row=1,column=i).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for i in range(17,21):
            worksheet_reassign.cell(row=1,column=i).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            no_phone_sheet.cell(row=1,column=i).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")      

        
    return output_filepath


def insert_reassign_columns(reassign_df):
    length = reassign_df.shape[0]
    reassign_df['New Investment Owner'] = ['']*length
    reassign_df['New Investment Owner ID'] = ['']*length
    reassign_df['Account Name with LEAD'] = reassign_df['Account Name'].apply(process_column)
    reassign_df['Matchcode'] = reassign_df['Account ID'].astype(str) + "-" + reassign_df['Account Name with LEAD'].astype(str)

    check = ['']*length
    attended = ['']*length
    future = ['']*length
    past = ['']*length
    fs = ['']*length
    contactid = ['']*length
    
    reassign_df['Check Webinar Reg'] = check
    reassign_df['Attended?'] = attended
    reassign_df['Future Webinar?'] = future
    reassign_df['Past Webinar?'] = past
    reassign_df['Force Sync'] = fs
    reassign_df['Contact ID'] = contactid


    col_names = ['New Investment Owner',	'New Investment Owner ID',	'Matchcode',	'Account Name with LEAD', 'Force Sync', 'Contact ID', 'Created Date',	'Account ID',	'Account Name',	'Confirmed Email Addresses',	'Call Notes',\
                 	'How much are you looking to invest?',	'How much are you looking to invest? (WR)',	'How much are you looking to invest? (EG)',	'Are you an accredited investor?',	'Lead Score',	\
                        'Check Webinar Reg',	'Attended?',	'Future Webinar?',	'Past Webinar?',	'Tags',	'Mailing State 1',	'Webinar Attendance',	'Investment Owner',	'Last Modified By',	\
                            'Created By',	'HNW indicator',	'Investment Disposition']
    reassign_df = reassign_df[col_names]
    reassign_df['How much are you looking to invest? (WR)'].replace(np.nan,0, inplace=True)
    reassign_df['How much are you looking to invest? (EG)'].replace(np.nan,0, inplace=True)
    reassign_df['How much are you looking to invest?'].replace(np.nan,0, inplace=True)

    reassign_df['How much are you looking to invest? (EG)'] = reassign_df['How much are you looking to invest? (EG)'].astype(str)
    reassign_df['How much are you looking to invest? (WR)'] = reassign_df['How much are you looking to invest? (WR)'].astype(str)

    reassign_df['sortWR'] = reassign_df['How much are you looking to invest? (WR)'].str.extract(r"(\d+)", expand=False).astype(float)
    reassign_df['sortWR'].replace(np.nan,0, inplace=True)
    reassign_df['sortEG'] = reassign_df['How much are you looking to invest? (EG)'].str.extract(r"(\d+)", expand=False).astype(float)
    reassign_df['sortEG'].replace(np.nan,0, inplace=True)
    reassign_df['sort'] = reassign_df['How much are you looking to invest?'].str.extract(r"(\d+)", expand=False).astype(float)
    reassign_df['sort'].replace(np.nan,0, inplace=True)


    reassign_df.sort_values(by=['sortWR', 'sortEG', 'sort'],axis=0, inplace=True, ascending=False)
    reassign_df = reassign_df.drop('sortWR', axis=1)
    reassign_df = reassign_df.drop('sortEG', axis=1)
    reassign_df = reassign_df.drop('sort', axis=1)



    reassign_df = reassign_df[reassign_df['Tags'] != 'webinar']
    reassign_df.dropna(subset=['Tags'], inplace=True)


    return reassign_df

def process_column(value):
    if ' - LEAD' in value:
        return value
    elif '- LEAD' in value:
        return value
    elif '- Lead' in value:
        return value
    elif ' - lead' in value:
        return value
    elif '- lead' in value:
        return value
    else:
        return value + ' - LEAD'

def acct_name_with_LEAD_faster(df):
    
    # path = os.getcwd()+'\\'+filename
    # contacts_sheet = pd.read_excel(filepath, sheet_name='Investor Accounts with Contacts', na_values='')
    # investor_accts_contacts = pd.DataFrame(contacts_sheet)
    # tqdm.pandas()
    # progress_bar = window["progressbar"]

    # Apply your function to the DataFrame series
    tqdm.pandas(desc="Processing")

    df['Account Name with LEAD'] = df['Account Name'].apply(process_column)
    df['Matchcode'] = df['Account ID'].astype(str) + "-" + df['Account Name with LEAD'].astype(str)
    col_names = ['Matchcode','Account Name with LEAD', 'Account ID', 'Account Name', 'Investment Owner', 'Investment Disposition', 'Confirmed Email Addresses', 'Last Name', 'Phone', 'Email', 'Force Sync', 'Contact ID']
    df = df[col_names]

    df['Phone'].replace('', np.nan, inplace=True)
    df['Phone'].replace('Unknown', np.nan, inplace=True)
    df['Phone'].replace('0', np.nan, inplace=True)
    df['Phone'].replace(0, np.nan, inplace=True)
    df['Phone'].replace('US', np.nan, inplace=True)
    df.dropna(subset=['Phone'], inplace=True)

    return df

def past_future_webinar(activities):
    todays_date = str(date.today())

    activities['Session Date'] = pd.to_datetime(activities['Session Date'])
    past_webinar = activities.loc[activities['Session Date'] < todays_date]
    future_webinar = activities.loc[activities['Session Date'] >= todays_date]

    return (past_webinar, future_webinar)





